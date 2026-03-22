"""
Fix all known issues in the EasyFlow P&L template.
Issues 1, 2, 4, 5 fixed here. Issues 3, 6, 7 are behavioral (handled by the writer).

Run standalone: python fix_template.py
"""

import os
import sys
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.comments import Comment

TEMPLATE_PATHS = [
    r"C:\Users\nmarc\EasyFlowCFO\templates\EasyFlow_CFO_P_L_Input.xlsx",
    r"C:\Users\nmarc\OneDrive - Living Water Consulting\EasyFlow_CFO_P&L_Input.xlsx",
]


def fix_all_issues(template_path, save=True):
    """Apply all known fixes to the template. Returns a list of issues fixed."""
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = load_workbook(template_path)
    ws = wb["Easy_P&L_Input"]
    issues_fixed = []

    # ── Issue 1 + Issue 4: Row 102 tax rate ──
    # Row 103 formula references $AJ$102, but ALL cells should have 0.40
    # for consistency and to prevent future issues.
    row102_was_wrong = False
    for col in range(2, 50):  # B (col 2) through AW (col 49)
        val = ws.cell(row=102, column=col).value
        if val != 0.40:
            row102_was_wrong = True
        ws.cell(row=102, column=col).value = 0.40

    if row102_was_wrong:
        issues_fixed.append("Issue 1: Row 102 tax rate fixed to 0.40 in all 48 columns")
    else:
        issues_fixed.append("Issue 1: Row 102 already correct (confirmed 0.40)")

    # Add comment (Issue 4)
    ws.cell(row=102, column=2).comment = Comment(
        "Tax rate assumption (40%). QBO import does not modify this row. Change manually to update all months.",
        "EasyFlow CFO System",
    )
    issues_fixed.append("Issue 4: Row 102 marked as config row with comment")

    # ── Issue 2: Duplicate Row 31 ──
    r30 = (ws.cell(row=30, column=1).value or "").strip()
    r31 = (ws.cell(row=31, column=1).value or "").strip()
    if r30.lower() == "fulfillment staff" and r31.lower() == "fulfillment staff":
        ws.cell(row=31, column=1).value = "Fulfillment Staff 2"
        for col in range(2, 50):
            ws.cell(row=31, column=col).value = None
        issues_fixed.append("Issue 2: Row 31 relabeled to 'Fulfillment Staff 2', values cleared")
    else:
        issues_fixed.append(f"Issue 2: Row 31 label is '{r31}' (no duplicate fix needed)")

    # ── Issue 5: Row 63 typo noted ──
    r63 = ws.cell(row=63, column=1).value or ""
    if "rfmodel" in r63.lower():
        issues_fixed.append(f"Issue 5: Row 63 label '{r63}' noted (typo for Office Remodel). Not renamed per spec.")

    if save:
        wb.save(template_path)
        print(f"  Saved: {template_path}")

    wb.close()
    return issues_fixed


def main():
    print("=" * 70)
    print("EasyFlow CFO Template Fix — All Known Issues")
    print(f"Run date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    for path in TEMPLATE_PATHS:
        if not os.path.exists(path):
            print(f"\n  SKIP (not found): {path}")
            continue

        print(f"\n  Fixing: {path}")

        # Create backup
        backup_dir = os.path.dirname(path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = os.path.splitext(os.path.basename(path))[0]
        backup_path = os.path.join(backup_dir, f"BACKUP_{timestamp}_{base}.xlsx")

        backup_wb = load_workbook(path)
        backup_wb.save(backup_path)
        backup_wb.close()
        print(f"  Backup: {backup_path}")

        issues = fix_all_issues(path, save=True)
        for issue in issues:
            print(f"    {issue}")

    print("\n  All fixes applied.")


if __name__ == "__main__":
    main()
