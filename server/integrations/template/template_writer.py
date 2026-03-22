"""
QBO to Easy Numbers P&L Template Filler.

Reads QBO data (API JSON or Excel export), aggregates by month,
and writes correct dollar amounts into the correct rows of
EasyFlow_CFO_P_L_Input.xlsx.

The template structure is fixed. Never overwrite a calculated row.
"""

import math
import os
import re
from collections import defaultdict
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.comments import Comment

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CONSTANTS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

INPUT_ROWS = {
    5:  'Organizing',
    6:  'Cleaning',
    7:  'Consulting Income',
    8:  'Workshops',
    9:  'Recurring Revenue/Subscription',
    10: 'Interest Income',
    11: 'Other',
    15: 'Subcontractors',
    16: 'Freight/Shipping',
    17: 'Project Materials / Supplies',
    18: 'Delivery & Transportation',
    19: 'Delivery Mileage & Tolls',
    20: 'Meals & Entertainment',
    21: 'Stripe Fees',
    22: 'Merchant account fees',
    29: 'Wages - [Owner]',
    30: 'Fulfillment Staff',
    32: 'Cleaning Fulfillment Staff',
    33: 'Client Success (delivery-only)',
    42: 'Commission on Sales',
    43: 'Referrals / Affiliate Commissions',
    44: 'Website Ads',
    45: 'Social Media Ads',
    46: 'Marketing Agency/Video editing',
    47: 'Branding (Logos, Design, Brand Assets)',
    48: 'Graphic Design (customer-facing)',
    49: 'Business Cards (Marketing)',
    50: 'CRM / Email Tools',
    51: 'Event Sponsorships',
    52: 'Website Design',
    53: 'Legacy Advertising',
    54: 'Trade Shows/ Promo Material',
    55: 'Gifts',
    62: 'Office Rent',
    63: 'Office Rfmodel',
    64: 'Utilities',
    65: 'Internet',
    66: 'Cleaning',
    70: 'Owner Payroll',
    76: 'FICA / FUTA / SUTA',
    77: 'Health Insurance',
    78: '401k Match',
    79: 'HSA Contributions',
    83: 'Entertainment meals',
    84: 'Supplies & Office',
    85: 'Shipping & Postage',
    86: 'Insurance',
    87: 'Accounting',
    88: 'Legal',
    89: 'Software and Subscriptions',
    90: 'Business Coaching & Education',
    91: 'Bank Fees & Service Charges',
    92: 'Memberships',
    93: 'Gifts',
    94: 'Charity',
}

CALCULATED_ROWS = frozenset([
    12, 23, 25, 26, 34, 36, 38, 39,
    56, 58, 67, 71, 73, 80, 95, 96,
    98, 99, 100, 103, 106, 107,
])

CONFIG_ROWS = frozenset([102])

SECTION_HEADER_ROWS = frozenset([4, 14, 28, 41, 60, 61, 69, 75, 82])


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# COLUMN INDEX
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def col_index(year, month):
    """
    Return the 1-based column index for a given year/month.
    Jan 2023 = 1 (column B), Dec 2026 = 48 (column AW).
    Raises ValueError if outside the 2023-2026 range.
    """
    if year < 2023 or year > 2026:
        raise ValueError(f"Year {year} is outside the template range 2023-2026")
    if month < 1 or month > 12:
        raise ValueError(f"Month {month} is not valid (1-12)")
    idx = ((year - 2023) * 12) + (month - 1) + 1
    if idx < 1 or idx > 48:
        raise ValueError(f"Column index {idx} is outside template range 1-48")
    return idx


def excel_col(year, month):
    """Return 1-based Excel column number (A=1, B=2). For openpyxl."""
    return col_index(year, month) + 1


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ROW GUARDS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class TemplateRowError(Exception):
    pass


def validate_write(row, col_idx, value):
    """
    Validate that a write is allowed. Raises on violation.
    col_idx is 1-based column index (B=2 in Excel).
    """
    if row in CALCULATED_ROWS:
        raise TemplateRowError(
            f"Row {row} is a calculated row ({INPUT_ROWS.get(row, 'formula')}). Cannot write QBO data here."
        )
    if row in CONFIG_ROWS:
        return 'skip'
    if row in SECTION_HEADER_ROWS:
        return 'skip'
    if row not in INPUT_ROWS:
        raise TemplateRowError(
            f"Row {row} is not a valid input row. Valid rows: {sorted(INPUT_ROWS.keys())}"
        )
    if col_idx < 2 or col_idx > 49:
        raise TemplateRowError(
            f"Column {col_idx} is outside template range (B=2 through AW=49)"
        )
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return 0
    return value


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# OWNER PAY SPLIT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

DEFAULT_SPLITS = {
    'solo_service': 0.70,
    'small_team':   0.50,
    'larger_team':  0.30,
    'retail_product': 0.40,
    'unknown':      0.50,
}


def owner_pay_split(total, direct_pct):
    """
    Split total owner pay into direct labor (Row 29) and management (Row 70).
    Uses floor + subtraction to guarantee exact sum.
    """
    if total <= 0:
        return 0, 0
    row29 = math.floor(total * direct_pct)
    row70 = total - row29
    return row29, row70


def get_default_split(business_type):
    """Return the default direct labor percentage for a business type."""
    return DEFAULT_SPLITS.get(business_type, 0.50)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# GRANULARITY DETECTION
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def detect_granularity(dates):
    """
    Detect whether a list of dates represents daily, weekly, monthly,
    quarterly, or annual data.
    """
    if not dates or len(dates) < 2:
        return 'unknown'
    sorted_dates = sorted(set(dates))
    if len(sorted_dates) < 2:
        return 'unknown'
    diffs = []
    for i in range(1, min(10, len(sorted_dates))):
        diffs.append((sorted_dates[i] - sorted_dates[i - 1]).days)
    avg_diff = sum(diffs) / len(diffs)
    if avg_diff <= 8:
        return 'daily_or_weekly'
    if avg_diff <= 35:
        return 'monthly'
    if avg_diff <= 95:
        return 'quarterly'
    return 'annual'


def aggregate_to_monthly(transactions):
    """
    Aggregate transaction-level data to monthly totals.
    Input: list of {account_name, date, amount, row (optional)}
    Output: list of {account_name, year, month, amount, row (optional)}
    """
    buckets = defaultdict(lambda: defaultdict(float))
    row_map = {}

    for txn in transactions:
        d = txn['date']
        if isinstance(d, str):
            d = datetime.strptime(d[:10], '%Y-%m-%d').date()
        elif isinstance(d, datetime):
            d = d.date()
        key = (txn['account_name'], d.year, d.month)
        buckets[txn['account_name']][(d.year, d.month)] += txn['amount']
        if 'row' in txn:
            row_map[txn['account_name']] = txn['row']

    result = []
    for name, periods in buckets.items():
        for (year, month), amount in sorted(periods.items()):
            entry = {'account_name': name, 'year': year, 'month': month, 'amount': amount}
            if name in row_map:
                entry['row'] = row_map[name]
            result.append(entry)
    return result


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# QBO ACCOUNT → TEMPLATE ROW MAPPING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _norm(s):
    return re.sub(r'\s+', ' ', s.strip().lower())


_EXACT_LABEL_MAP = {}
for _r, _label in INPUT_ROWS.items():
    _EXACT_LABEL_MAP[_norm(_label)] = _r

# Substring-based rules: (substring_pattern, target_row, section_constraint)
_REVENUE_RULES = [
    ('organizing', 5, None),
    ('consulting', 7, None),
    ('workshop', 8, None),
    ('training', 8, None),
    ('speaking', 8, None),
    ('recurring', 9, None),
    ('subscription', 9, None),
    ('retainer', 9, None),
    ('interest', 10, None),
]

_COGS_RULES = [
    ('subcontractor', 15, None),
    ('contract labor', 15, None),
    ('1099', 15, None),
    ('freight', 16, 'cogs'),
    ('shipping', 16, 'cogs'),
    ('material', 17, None),
    ('supply', 17, 'cogs'),
    ('supplies', 17, 'cogs'),
    ('parts', 17, None),
    ('inventory', 17, None),
    ('job material', 17, None),
    ('delivery', 18, 'cogs'),
    ('transportation', 18, 'cogs'),
    ('mileage', 19, None),
    ('toll', 19, None),
    ('meals', 20, 'cogs'),
    ('entertainment', 20, 'cogs'),
    ('stripe', 21, None),
    ('merchant', 22, None),
    ('processing fee', 22, None),
    ('transaction fee', 22, None),
]

_LABOR_RULES = [
    ('fulfillment', 30, None),
    ('field staff', 30, None),
    ('production staff', 30, None),
    ('cleaning staff', 32, None),
    ('cleaning labor', 32, None),
    ('client success', 33, None),
    ('client services', 33, None),
]

_MARKETING_RULES = [
    ('advertising', 53, None),
    ('commission', 42, None),
    ('referral', 43, None),
    ('affiliate', 43, None),
    ('website ad', 44, None),
    ('google ad', 44, None),
    ('bing', 44, None),
    ('social media', 45, None),
    ('facebook', 45, None),
    ('instagram', 45, None),
    ('tiktok', 45, None),
    ('marketing agency', 46, None),
    ('video edit', 46, None),
    ('video', 46, 'marketing'),
    ('brand', 47, 'marketing'),
    ('logo', 47, None),
    ('graphic design', 48, None),
    ('design', 48, 'marketing'),
    ('business card', 49, None),
    ('print marketing', 49, None),
    ('crm', 50, None),
    ('email marketing', 50, None),
    ('email tool', 50, None),
    ('sponsorship', 51, None),
    ('event sponsor', 51, None),
    ('website design', 52, None),
    ('website dev', 52, None),
    ('legacy ad', 53, None),
    ('print ad', 53, None),
    ('trade show', 54, None),
    ('promo material', 54, None),
    ('marketing', 53, None),
    ('ads', 53, None),
]

_OPEX_RULES = [
    ('rent', 62, None),
    ('lease', 62, None),
    ('remodel', 63, None),
    ('renovation', 63, None),
    ('improvement', 63, None),
    ('utilities', 64, None),
    ('electric', 64, None),
    ('water', 64, 'opex'),
    ('gas', 64, 'opex'),
    ('internet', 65, None),
    ('cable', 65, None),
    ('fica', 76, None),
    ('futa', 76, None),
    ('suta', 76, None),
    ('payroll tax', 76, None),
    ('health insurance', 77, None),
    ('medical', 77, 'opex'),
    ('401k', 78, None),
    ('retirement', 78, None),
    ('pension', 78, None),
    ('hsa', 79, None),
    ('office supply', 84, None),
    ('shipping', 85, 'opex'),
    ('postage', 85, None),
    ('insurance', 86, None),
    ('accounting', 87, None),
    ('bookkeeping', 87, None),
    ('cpa', 87, None),
    ('legal', 88, None),
    ('attorney', 88, None),
    ('software', 89, None),
    ('subscription', 89, 'opex'),
    ('saas', 89, None),
    ('coaching', 90, None),
    ('education', 90, None),
    ('bank fee', 91, None),
    ('wire fee', 91, None),
    ('service charge', 91, None),
    ('membership', 92, None),
    ('dues', 92, None),
    ('charity', 94, None),
    ('donation', 94, None),
]

# Owner pay keywords
_OWNER_PAY_KEYWORDS = [
    'officer compensation', 'officer salary', 'officer wages',
    'owner compensation', 'owner salary', 'owner wages', 'owner draw',
    'guaranteed payments', 'partner compensation', 'member compensation',
    's-corp owner wages', 'shareholder wages',
]


def map_account_to_row(account_name, section='', qbo_type=''):
    """
    Map a QBO account name to a template row number.
    Returns (row_number, confidence, flags) or (None, 0, flags) if unmatched.

    section: 'income', 'cogs', 'expense', 'other_expense'
    """
    name = _norm(account_name)
    flags = []

    # Step 1: exact label match
    if name in _EXACT_LABEL_MAP:
        row = _EXACT_LABEL_MAP[name]
        if row == 10:
            flags.append('non_operating_income')
        return row, 1.0, flags

    sect = section.lower() if section else ''

    # Owner pay detection (any section)
    for kw in _OWNER_PAY_KEYWORDS:
        if kw in name:
            if sect in ('cogs', 'costofgoodssold', 'direct labor'):
                return 29, 0.90, ['owner_direct_labor']
            else:
                return 70, 0.95, ['owner_management']

    # "cleaning" disambiguation: income section → row 6, opex → row 66
    if 'cleaning' in name:
        if sect in ('income', 'revenue'):
            return 6, 0.90, []
        elif sect in ('expense', 'other_expense', 'opex', 'operating'):
            return 66, 0.90, []

    # "gift" disambiguation: marketing section → 55, opex → 93
    if 'gift' in name:
        if sect in ('marketing',):
            return 55, 0.85, []
        else:
            return 93, 0.85, []

    # "entertainment" disambiguation
    if 'entertainment' in name:
        if sect in ('cogs', 'costofgoodssold'):
            return 20, 0.85, ['confirm_client_facing']
        else:
            return 83, 0.85, []

    # "supply" or "supplies" disambiguation
    if ('supply' in name or 'supplies' in name) and sect not in ('cogs', 'costofgoodssold'):
        return 84, 0.85, []

    # Step 2: rule-based substring match
    all_rules = _REVENUE_RULES + _COGS_RULES + _LABOR_RULES + _MARKETING_RULES + _OPEX_RULES
    for pattern, target_row, section_constraint in all_rules:
        if pattern in name:
            if section_constraint is None:
                if target_row == 10:
                    flags.append('non_operating_income')
                return target_row, 0.85, flags
            elif section_constraint == 'cogs' and sect in ('cogs', 'costofgoodssold'):
                return target_row, 0.85, flags
            elif section_constraint == 'marketing' and sect in ('marketing',):
                return target_row, 0.85, flags
            elif section_constraint == 'opex' and sect in ('expense', 'other_expense', 'opex', 'operating', 'expenses', 'otherexpenses'):
                return target_row, 0.85, flags

    # Generic revenue fallback for income section
    if sect in ('income', 'revenue', 'otherincome'):
        if 'sales' in name or 'revenue' in name or 'income' in name:
            return 11, 0.80, []
        return None, 0, []

    # W-2 wages in COGS/labor section
    labor_kw = ('wages', 'salary', 'salaries', 'payroll', 'labor', 'labour', 'crew', 'technician')
    if sect in ('cogs', 'costofgoodssold', 'direct labor') and any(k in name for k in labor_kw):
        return 30, 0.80, []

    # Expense section fallthrough → row 94 (Charity/catch-all)
    if sect in ('expense', 'other_expense', 'opex', 'operating', 'expenses', 'otherexpenses'):
        return 94, 0.70, ['unmapped_fallthrough']

    return None, 0, []


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# WRITER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def write_to_template(
    template_path,
    mapped_data,
    overwrite_existing=False,
    fix_known_issues=True,
    output_path=None,
):
    """
    Write mapped QBO data to the Easy Numbers P&L template.

    mapped_data: list of dicts with keys:
        row (int): template row number
        col_index (int): 1-based month index (1=Jan 2023, 48=Dec 2026)
        value (float): dollar amount

    Returns a write_summary dict.
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = load_workbook(template_path)
    ws = wb['Easy_P&L_Input']

    issues_fixed = []

    if fix_known_issues:
        from fix_template import fix_all_issues as _fix
        # We re-apply fixes inline rather than importing to keep self-contained
        # Fix Row 102
        for col in range(2, 50):
            ws.cell(row=102, column=col).value = 0.40
        ws.cell(row=102, column=2).comment = Comment(
            "Tax rate assumption (40%). QBO import does not modify this row.",
            "EasyFlow CFO System",
        )
        issues_fixed.append("Row 102 tax rate set to 0.40")

        # Fix Row 31
        r30 = (ws.cell(row=30, column=1).value or '').strip()
        r31 = (ws.cell(row=31, column=1).value or '').strip()
        if r30.lower() == 'fulfillment staff' and r31.lower() in ('fulfillment staff', 'fulfillment staff 2'):
            if r31.lower() != 'fulfillment staff 2':
                ws.cell(row=31, column=1).value = 'Fulfillment Staff 2'
                for col in range(2, 50):
                    ws.cell(row=31, column=col).value = None
                issues_fixed.append("Row 31 relabeled to Fulfillment Staff 2")

    cells_written = 0
    cells_skipped = 0
    skipped_log = []
    months_populated = set()
    flags = []

    for entry in mapped_data:
        row = entry['row']
        ci = entry['col_index']
        value = entry['value']
        excel_column = ci + 1  # col_index 1 → Excel column B (2)

        result = validate_write(row, excel_column, value)
        if result == 'skip':
            cells_skipped += 1
            continue
        if isinstance(result, (int, float)) and result == 0 and value is None:
            value = 0
        elif isinstance(result, (int, float)):
            value = result if result != 0 else value

        existing = ws.cell(row=row, column=excel_column).value
        if existing and existing != 0 and not overwrite_existing:
            cells_skipped += 1
            skipped_log.append({
                'row': row, 'col': excel_column,
                'existing': existing, 'proposed': value,
            })
            continue

        ws.cell(row=row, column=excel_column).value = value
        cells_written += 1

        # Track month
        year = 2023 + (ci - 1) // 12
        month = ((ci - 1) % 12) + 1
        months_populated.add(f"{year}-{month:02d}")

    # Row 10 flag
    for ci_check in range(1, 49):
        v = ws.cell(row=10, column=ci_check + 1).value
        if v and v != 0:
            flags.append('Interest Income (Row 10) is recorded but excluded from Profit Score operating revenue.')
            break

    # Owner pay validation
    for ci_check in range(1, 49):
        r29 = ws.cell(row=29, column=ci_check + 1).value or 0
        r70 = ws.cell(row=70, column=ci_check + 1).value or 0
        if isinstance(r29, (int, float)) and isinstance(r70, (int, float)):
            # Penny adjustment if needed
            pass

    sorted_months = sorted(months_populated) if months_populated else []

    if output_path is None:
        base = os.path.splitext(os.path.basename(template_path))[0]
        today = date.today().isoformat()
        output_path = os.path.join(os.path.dirname(template_path), f"{base}_{today}.xlsx")

    wb.save(output_path)
    wb.close()

    return {
        'cells_written': cells_written,
        'cells_skipped': cells_skipped,
        'skipped_log': skipped_log,
        'months_populated': sorted_months,
        'first_month': sorted_months[0] if sorted_months else None,
        'last_month': sorted_months[-1] if sorted_months else None,
        'issues_fixed': issues_fixed,
        'flags': flags,
        'output_path': output_path,
    }
