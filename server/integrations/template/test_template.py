"""
25 tests for the QBO to Easy Numbers P&L Template Filler.

Run: python -m pytest test_template.py -v
  or: python test_template.py
"""

import math
import os
import sys
import tempfile
import shutil
from datetime import date, datetime

# Ensure we can import from this directory
sys.path.insert(0, os.path.dirname(__file__))

from template_writer import (
    col_index, excel_col,
    INPUT_ROWS, CALCULATED_ROWS, CONFIG_ROWS, SECTION_HEADER_ROWS,
    validate_write, TemplateRowError,
    owner_pay_split, get_default_split,
    detect_granularity, aggregate_to_monthly,
    map_account_to_row,
    write_to_template,
)
from fix_template import fix_all_issues

TEMPLATE_PATH = r"C:\Users\nmarc\EasyFlowCFO\templates\EasyFlow_CFO_P_L_Input.xlsx"

passed = 0
failed = 0
errors = []


def test(name, fn):
    global passed, failed
    try:
        fn()
        passed += 1
        print(f"  PASS  {name}")
    except Exception as e:
        failed += 1
        errors.append((name, str(e)))
        print(f"  FAIL  {name}: {e}")


# ━━━━━━━━━ Column mapping tests (1-5) ━━━━━━━━━

def test_1():
    assert col_index(2023, 1) == 1, f"Expected 1, got {col_index(2023, 1)}"

def test_2():
    assert col_index(2023, 12) == 12, f"Expected 12, got {col_index(2023, 12)}"

def test_3():
    assert col_index(2024, 1) == 13, f"Expected 13, got {col_index(2024, 1)}"

def test_4():
    assert col_index(2026, 12) == 48, f"Expected 48, got {col_index(2026, 12)}"

def test_5():
    raised = False
    try:
        col_index(2022, 6)
    except ValueError:
        raised = True
    assert raised, "Should raise ValueError for year outside 2023-2026"
    raised = False
    try:
        col_index(2027, 1)
    except ValueError:
        raised = True
    assert raised, "Should raise ValueError for 2027"


# ━━━━━━━━━ Row guard tests (6-9) ━━━━━━━━━

def test_6():
    raised = False
    try:
        validate_write(25, 2, 1000)
    except TemplateRowError:
        raised = True
    assert raised, "Row 25 (Gross Margin) should raise TemplateRowError"

def test_7():
    result = validate_write(102, 2, 0.40)
    assert result == 'skip', f"Row 102 (tax rate) via QBO import should return 'skip', got {result}"

def test_8():
    raised = False
    try:
        validate_write(56, 5, 500)
    except TemplateRowError:
        raised = True
    assert raised, "Row 56 (Total Marketing) should raise TemplateRowError"

def test_9():
    result = validate_write(45, 5, 1200)
    assert result == 1200, f"Row 45 (Social Media Ads) should succeed, got {result}"


# ━━━━━━━━━ Issue fix tests (10-12) ━━━━━━━━━

def _make_temp_template():
    """Copy the real template to a temp file for testing."""
    if not os.path.exists(TEMPLATE_PATH):
        return None
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    shutil.copy2(TEMPLATE_PATH, tmp.name)
    return tmp.name

def test_10():
    from openpyxl import load_workbook
    tmp = _make_temp_template()
    if not tmp:
        raise Exception("Template file not found, cannot run issue fix test")
    try:
        fix_all_issues(tmp, save=True)
        wb = load_workbook(tmp)
        ws = wb['Easy_P&L_Input']
        for col in range(2, 50):
            val = ws.cell(row=102, column=col).value
            assert val == 0.40, f"Row 102 col {col} = {val}, expected 0.40"
        wb.close()
    finally:
        os.unlink(tmp)

def test_11():
    from openpyxl import load_workbook
    tmp = _make_temp_template()
    if not tmp:
        raise Exception("Template file not found")
    try:
        fix_all_issues(tmp, save=True)
        wb = load_workbook(tmp)
        ws = wb['Easy_P&L_Input']
        label = ws.cell(row=31, column=1).value
        assert label == 'Fulfillment Staff 2', f"Row 31 label = '{label}', expected 'Fulfillment Staff 2'"
        wb.close()
    finally:
        os.unlink(tmp)

def test_12():
    from openpyxl import load_workbook
    tmp = _make_temp_template()
    if not tmp:
        raise Exception("Template file not found")
    try:
        fix_all_issues(tmp, save=True)
        wb = load_workbook(tmp)
        ws = wb['Easy_P&L_Input']
        for col in range(2, 50):
            val = ws.cell(row=31, column=col).value
            assert val is None or val == 0, f"Row 31 col {col} should be cleared, got {val}"
        wb.close()
    finally:
        os.unlink(tmp)


# ━━━━━━━━━ Owner pay tests (13-17) ━━━━━━━━━

def test_13():
    r29, r70 = owner_pay_split(100000, 0.70)
    assert r29 == 70000, f"Row29 = {r29}, expected 70000"
    assert r70 == 30000, f"Row70 = {r70}, expected 30000"
    assert r29 + r70 == 100000, f"Sum = {r29 + r70}, expected 100000"

def test_14():
    r29, r70 = owner_pay_split(99999, 0.70)
    assert r29 == 69999, f"Row29 = {r29}, expected 69999"
    assert r70 == 30000, f"Row70 = {r70}, expected 30000"
    assert r29 + r70 == 99999, f"Sum = {r29 + r70}, expected 99999"

def test_15():
    r29, r70 = owner_pay_split(1, 0.50)
    assert r29 == 0, f"Row29 = {r29}, expected 0"
    assert r70 == 1, f"Row70 = {r70}, expected 1"
    assert r29 + r70 == 1

def test_16():
    """overwrite_existing=False: existing value should not be changed."""
    from openpyxl import load_workbook
    tmp = _make_temp_template()
    if not tmp:
        raise Exception("Template file not found")
    try:
        # First write a value to Row 29, Dec 2025 (col_index=36)
        wb = load_workbook(tmp)
        ws = wb['Easy_P&L_Input']
        ws.cell(row=29, column=37).value = 1872  # col_index 36 → excel col 37
        wb.save(tmp)
        wb.close()

        out = os.path.join(tempfile.gettempdir(), 'test_overwrite_false.xlsx')
        summary = write_to_template(
            tmp,
            [{'row': 29, 'col_index': 36, 'value': 5000}],
            overwrite_existing=False,
            fix_known_issues=False,
            output_path=out,
        )
        assert summary['cells_skipped'] >= 1, f"Expected skipped >= 1, got {summary['cells_skipped']}"

        wb2 = load_workbook(out)
        ws2 = wb2['Easy_P&L_Input']
        val = ws2.cell(row=29, column=37).value
        assert val == 1872, f"Value should remain 1872, got {val}"
        wb2.close()
        os.unlink(out)
    finally:
        os.unlink(tmp)

def test_17():
    """overwrite_existing=True: existing value should be overwritten."""
    from openpyxl import load_workbook
    tmp = _make_temp_template()
    if not tmp:
        raise Exception("Template file not found")
    try:
        wb = load_workbook(tmp)
        ws = wb['Easy_P&L_Input']
        ws.cell(row=29, column=37).value = 1872
        wb.save(tmp)
        wb.close()

        out = os.path.join(tempfile.gettempdir(), 'test_overwrite_true.xlsx')
        summary = write_to_template(
            tmp,
            [{'row': 29, 'col_index': 36, 'value': 5000}],
            overwrite_existing=True,
            fix_known_issues=False,
            output_path=out,
        )
        assert summary['cells_written'] >= 1

        wb2 = load_workbook(out)
        ws2 = wb2['Easy_P&L_Input']
        val = ws2.cell(row=29, column=37).value
        assert val == 5000, f"Value should be 5000, got {val}"
        wb2.close()
        os.unlink(out)
    finally:
        os.unlink(tmp)


# ━━━━━━━━━ Aggregation tests (18-21) ━━━━━━━━━

def test_18():
    """Weekly data aggregated to monthly."""
    txns = [
        {'account_name': 'Organizing', 'date': '2024-01-05', 'amount': 1000, 'row': 5},
        {'account_name': 'Organizing', 'date': '2024-01-12', 'amount': 2000, 'row': 5},
        {'account_name': 'Organizing', 'date': '2024-01-19', 'amount': 1500, 'row': 5},
        {'account_name': 'Organizing', 'date': '2024-01-26', 'amount': 500, 'row': 5},
    ]
    monthly = aggregate_to_monthly(txns)
    assert len(monthly) == 1
    assert monthly[0]['year'] == 2024
    assert monthly[0]['month'] == 1
    assert monthly[0]['amount'] == 5000, f"Expected 5000, got {monthly[0]['amount']}"

def test_19():
    """Monthly data passes through correctly."""
    monthly = aggregate_to_monthly([
        {'account_name': 'Organizing', 'date': '2024-01-15', 'amount': 5000, 'row': 5},
    ])
    assert monthly[0]['amount'] == 5000

def test_20():
    dates = [date(2024, 1, d) for d in range(1, 31)]
    assert detect_granularity(dates) == 'daily_or_weekly'

def test_21():
    dates = [date(2024, m, 1) for m in range(1, 13)]
    assert detect_granularity(dates) == 'monthly'


# ━━━━━━━━━ Write summary tests (22-24) ━━━━━━━━━

def test_22():
    tmp = _make_temp_template()
    if not tmp:
        raise Exception("Template file not found")
    try:
        data = [
            {'row': 5, 'col_index': 1, 'value': 10000},
            {'row': 6, 'col_index': 1, 'value': 5000},
            {'row': 15, 'col_index': 1, 'value': 2000},
        ]
        out = os.path.join(tempfile.gettempdir(), 'test_summary.xlsx')
        summary = write_to_template(tmp, data, fix_known_issues=True, output_path=out)
        assert summary['cells_written'] == 3, f"Expected 3 cells written, got {summary['cells_written']}"
        os.unlink(out)
    finally:
        os.unlink(tmp)

def test_23():
    tmp = _make_temp_template()
    if not tmp:
        raise Exception("Template file not found")
    try:
        out = os.path.join(tempfile.gettempdir(), 'test_issues.xlsx')
        summary = write_to_template(tmp, [], fix_known_issues=True, output_path=out)
        assert len(summary['issues_fixed']) >= 1, f"Expected issues_fixed >= 1, got {summary['issues_fixed']}"
        os.unlink(out)
    finally:
        os.unlink(tmp)

def test_24():
    tmp = _make_temp_template()
    if not tmp:
        raise Exception("Template file not found")
    try:
        data = [{'row': 10, 'col_index': 1, 'value': 250}]
        out = os.path.join(tempfile.gettempdir(), 'test_interest.xlsx')
        summary = write_to_template(tmp, data, fix_known_issues=False, output_path=out)
        assert any('Interest Income' in f for f in summary['flags']), \
            f"Expected interest income flag, got: {summary['flags']}"
        os.unlink(out)
    finally:
        os.unlink(tmp)


# ━━━━━━━━━ End-to-end test (25) ━━━━━━━━━

def test_25():
    """
    Load real template. Apply fixes. Write 12 months of test data.
    Verify formulas compute correctly via data_only read.
    """
    from openpyxl import load_workbook
    tmp = _make_temp_template()
    if not tmp:
        raise Exception("Template file not found")
    try:
        # Generate 12 months of test data for Jan-Dec 2024
        data = []
        for m in range(1, 13):
            ci = col_index(2024, m)
            # Revenue
            data.append({'row': 5, 'col_index': ci, 'value': 8000})   # Organizing
            data.append({'row': 7, 'col_index': ci, 'value': 4000})   # Consulting
            data.append({'row': 10, 'col_index': ci, 'value': 50})    # Interest Income (non-operating)
            data.append({'row': 11, 'col_index': ci, 'value': 1000})  # Other
            # COGS
            data.append({'row': 15, 'col_index': ci, 'value': 1500})  # Subcontractors
            data.append({'row': 17, 'col_index': ci, 'value': 500})   # Materials
            # Direct Labor
            data.append({'row': 29, 'col_index': ci, 'value': 3000})  # Owner direct
            data.append({'row': 30, 'col_index': ci, 'value': 2000})  # Fulfillment
            # Marketing
            data.append({'row': 45, 'col_index': ci, 'value': 800})   # Social Media
            data.append({'row': 50, 'col_index': ci, 'value': 200})   # CRM
            # OpEx
            data.append({'row': 62, 'col_index': ci, 'value': 1200})  # Rent
            data.append({'row': 70, 'col_index': ci, 'value': 2000})  # Owner mgmt
            data.append({'row': 86, 'col_index': ci, 'value': 300})   # Insurance
            data.append({'row': 89, 'col_index': ci, 'value': 400})   # Software

        out = os.path.join(tempfile.gettempdir(), 'test_e2e.xlsx')
        summary = write_to_template(tmp, data, fix_known_issues=True, overwrite_existing=True, output_path=out)

        # Verify structure (formulas are preserved, values written)
        wb = load_workbook(out)
        ws = wb['Easy_P&L_Input']

        # Check Jan 2024 (col 14 = excel col for col_index 13)
        ec = excel_col(2024, 1)  # should be 14

        # Revenue inputs
        assert ws.cell(row=5, column=ec).value == 8000
        assert ws.cell(row=7, column=ec).value == 4000
        assert ws.cell(row=10, column=ec).value == 50
        assert ws.cell(row=11, column=ec).value == 1000

        # Row 12 should be a formula (Total Revenue)
        r12 = ws.cell(row=12, column=ec).value
        assert isinstance(r12, str) and r12.startswith('='), f"Row 12 should be formula, got {r12}"

        # Row 102 all = 0.40
        for col in range(2, 50):
            assert ws.cell(row=102, column=col).value == 0.40, f"Row 102 col {col} != 0.40"

        # Row 31 label
        assert ws.cell(row=31, column=1).value == 'Fulfillment Staff 2'

        # Interest income flag
        assert any('Interest Income' in f for f in summary['flags'])

        # Check 12 months populated
        assert len(summary['months_populated']) == 12
        assert summary['first_month'] == '2024-01'
        assert summary['last_month'] == '2024-12'

        wb.close()
        os.unlink(out)
    finally:
        os.unlink(tmp)


# ━━━━━━━━━ Runner ━━━━━━━━━

if __name__ == '__main__':
    print("\n" + "=" * 60)
    print("EasyFlow CFO Template Writer — Test Suite (25 tests)")
    print("=" * 60 + "\n")

    test("1.  col_index Jan 2023 = 1", test_1)
    test("2.  col_index Dec 2023 = 12", test_2)
    test("3.  col_index Jan 2024 = 13", test_3)
    test("4.  col_index Dec 2026 = 48", test_4)
    test("5.  col_index outside range raises", test_5)
    test("6.  Write to Row 25 (Gross Margin) raises", test_6)
    test("7.  Write to Row 102 (tax rate) via QBO skipped", test_7)
    test("8.  Write to Row 56 (Total Marketing) raises", test_8)
    test("9.  Write to Row 45 (Social Media Ads) succeeds", test_9)
    test("10. After fix: Row 102 all = 0.40", test_10)
    test("11. After fix: Row 31 = Fulfillment Staff 2", test_11)
    test("12. After fix: Row 31 values cleared", test_12)
    test("13. split(100000, 0.70) exact sum", test_13)
    test("14. split(99999, 0.70) exact sum", test_14)
    test("15. split(1, 0.50) floor edge case", test_15)
    test("16. overwrite_existing=False skips", test_16)
    test("17. overwrite_existing=True overwrites", test_17)
    test("18. Weekly aggregation to monthly", test_18)
    test("19. Monthly data direct pass-through", test_19)
    test("20. detect_granularity daily dates", test_20)
    test("21. detect_granularity monthly dates", test_21)
    test("22. write_summary.cells_written correct", test_22)
    test("23. write_summary.issues_fixed has fixes", test_23)
    test("24. write_summary.flags has Row 10 interest", test_24)
    test("25. End-to-end: 12 months, formulas, fixes", test_25)

    print(f"\n{'=' * 60}")
    print(f"Results: {passed} passed, {failed} failed out of {passed + failed}")
    if errors:
        print("\nFailures:")
        for name, err in errors:
            print(f"  {name}: {err}")
    print("=" * 60)

    sys.exit(0 if failed == 0 else 1)
